import cv2
import os
import pytesseract
from matplotlib import pyplot as pt
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.borders import Border, Side
from drawingNum import GetString


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


wb = Workbook()


default_sheet = wb.active
wb.remove(default_sheet)


for image in range(1, 7):

    img_path = os.path.join("images", f'{image:02}.png')

    
    if not os.path.exists(img_path):
        print(f"Image {image:02}.png not found.")
        continue

    # Read image
    init_img = cv2.imread(img_path, 0)

    if init_img is None:
        print(f"Cannot read {image:02}.png")
        continue

    init_row, init_col = init_img.shape

  
    img = init_img[12:init_row-15, 12:init_col-12]

    nrow, ncol = img.shape

  
    _, bin_img = cv2.threshold(
        img,
        127,
        255,
        cv2.THRESH_BINARY_INV
    )

    horiz_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (1, ncol // 150)
    )

    eroded_verti = cv2.erode(
        bin_img,
        horiz_kernel,
        iterations=5
    )

    vertical_lines = cv2.dilate(
        eroded_verti,
        horiz_kernel,
        iterations=5
    )

  
    verti_kernel = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (nrow // 150, 1)
    )

    eroded_hori = cv2.erode(
        bin_img,
        verti_kernel,
        iterations=5
    )

    horizontal_lines = cv2.dilate(
        eroded_hori,
        verti_kernel,
        iterations=5
    )

    combined_lines = cv2.bitwise_or(
        vertical_lines,
        horizontal_lines
    )


    rect_kernel3 = cv2.getStructuringElement(
        cv2.MORPH_RECT,
        (3, 3)
    )

    drawingMask = cv2.erode(
        combined_lines,
        rect_kernel3,
        iterations=2
    )

    drawingMask = cv2.dilate(
        drawingMask,
        rect_kernel3,
        iterations=50
    )

    table_lines = drawingMask + np.bitwise_not(combined_lines)

    
    table_lines_dil = cv2.dilate(
        np.bitwise_not(table_lines),
        rect_kernel3,
        iterations=5
    )

    
    contours, hierarchy = cv2.findContours(
        table_lines_dil,
        cv2.RETR_TREE,
        cv2.CHAIN_APPROX_SIMPLE
    )

    sorted_contours = sorted(
        contours,
        key=cv2.contourArea,
        reverse=False
    )

    table_bgr = cv2.cvtColor(
        table_lines,
        cv2.COLOR_GRAY2BGR
    )

    for i in range(len(sorted_contours)):

        cntr = sorted_contours[i]

        x, y, w, h = cv2.boundingRect(cntr)

        if w < 30 or h < 30:

            cv2.drawContours(
                table_bgr,
                sorted_contours,
                i,
                (255, 255, 255),
                thickness=-1
            )


    table_only = cv2.cvtColor(
        table_bgr,
        cv2.COLOR_BGR2GRAY
    )

    _, table_only = cv2.threshold(
        table_only,
        150,
        255,
        cv2.THRESH_BINARY
    )

    table_only_copy = cv2.copyMakeBorder(
        table_only,
        5,
        5,
        5,
        5,
        cv2.BORDER_CONSTANT,
        0
    )

    table_lines_dil2 = cv2.dilate(
        np.bitwise_not(table_only_copy),
        rect_kernel3,
        iterations=1
    )

 
    cell_cntr, hierarchy = cv2.findContours(
        table_lines_dil2,
        cv2.RETR_TREE,
        cv2.CHAIN_APPROX_SIMPLE
    )

    table_bgr2 = cv2.cvtColor(
        table_only,
        cv2.COLOR_GRAY2BGR
    )

    keywords = [
        "DRAWING NUMBER",
        "DRAWING NO",
        "DRAWN BY",
        "DRAWN",
        "CHECKED BY",
        "CHECKED",
        "TITLE",
        "DRAWING TITLE",
        "APPROVED BY",
        "APPROVED",
        "CONTRACTOR",
        "COMPANY",
        "UNIT",
        "STATUS",
        "PAGE",
        "PROJECT NO",
        "PROJECT NUM",
        "LANG",
        "CAD NO",
        "FONT",
        "FONT STYLE",
        "AMENDMENTS"
    ]

    useful_cells = []

    
    for c in cell_cntr:

        coordinates = cv2.boundingRect(c)

        x, y, w, h = coordinates

        rect_area = w * h

        if rect_area < ((nrow // 4) * (ncol // 4)) and h < 400:

            cell = img[y:y+h, x:x+w]

            string = pytesseract.image_to_string(
                cell,
                config='--psm 6'
            ).strip()

            string_list = string.splitlines()

            for k in keywords:

                if k in string.upper():

                    cell_info = [
                        k,
                        coordinates,
                        string_list
                    ]

                    useful_cells.append(cell_info)

            cv2.rectangle(
                table_bgr2,
                (x, y),
                (x+w, y+h),
                (0, 0, 0),
                -1
            )


    table_mask = cv2.cvtColor(
        table_bgr2,
        cv2.COLOR_BGR2GRAY
    )

    table_mask = cv2.dilate(
        np.bitwise_not(table_mask),
        rect_kernel3,
        iterations=5
    )

    drawing = np.bitwise_not(bin_img) + table_mask

    drawing[drawing >= 5] = 255
    drawing[drawing < 5] = 0

    
    table_data = []

    for c in useful_cells:

        if c not in table_data:
            table_data.append(c)


    ws = wb.create_sheet(f'{image:02}.png')

    ws.append(["Field Title", "Content"])

    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    ws.cell(row=1, column=1).border = thick_border
    ws.cell(row=1, column=2).border = thick_border

   
    for info in table_data:

        value = ""

        if len(info[2]) > 1:
            value = info[2][1]

        ws.append([info[0], value])

    wb.save(filename='drawingInfo2.xlsx')


    writeFolder = "extracted"

    if not os.path.exists(writeFolder):
        os.makedirs(writeFolder)

   
    output_image_path = os.path.join(
        writeFolder,
        f'drawing{image:02}.png'
    )

    cv2.imwrite(output_image_path, drawing)

    print(f"Image {image:02}.png processed.")

print("\nFinished Successfully")